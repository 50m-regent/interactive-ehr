"""DWHテーブル設計書(xlsx)からPydanticモデルを自動生成するスクリプト.

Usage:
    uv run python scripts/generate_models.py
"""

from __future__ import annotations

import re
import subprocess
import sys
from dataclasses import dataclass, field
from datetime import date, time
from pathlib import Path

import openpyxl

# パス設定
PROJECT_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = PROJECT_ROOT / "data" / "dwh_table_design_2025-11-01.xlsx"
MODELS_DIR = PROJECT_ROOT / "src" / "interactive_ehr" / "models"

# SQL型 → Python型マッピング
SQL_TYPE_MAP: dict[str, str] = {
    "VARCHAR": "str",
    "SMALLINT": "int",
    "FLOAT": "float",
    "DECIMAL": "Decimal",
    "DATE": "date",
    "TIME": "time",
}

# グループ → ファイル名マッピング
# ORDER は大きいので3ファイルに分割
GROUP_FILE_MAP: dict[str, str] = {
    "PATIENT": "patient",
    "MR": "mr",
    "NURSE": "nurse",
    "OTHER": "other",
}

# ORDERグループのテーブル → サブファイル分類
ORDER_EXAM_TABLES = frozenset(
    {
        "検体検査",
        "検体検査結果",
        "細菌検査",
        "一般細菌塗抹鏡検結果",
        "一般細菌培養同定検査結果",
        "一般細菌感受性検査結果",
        "一般細菌追加検査結果",
        "抗酸菌塗抹鏡検結果",
        "抗酸菌培養検査結果",
        "抗酸菌遺伝子検査結果",
        "抗酸菌同定検査結果",
        "抗酸菌感受性検査結果",
        "その他細菌検査結果",
        "病理検査",
        "病理検査レポート",
        "生理検査",
        "生理検査レポート",
        "内視鏡検査",
        "内視鏡検査レポート",
        "放射線検査",
        "放射線検査オーダー",
        "放射線検査薬剤",
        "放射線検査材料",
        "放射線検査レポート",
        "RI検査",
        "RI検査レポート",
    }
)

ORDER_TREATMENT_TABLES = frozenset(
    {
        "処方",
        "処方指示",
        "処方実施",
        "注射指示",
        "注射実施",
        "処置",
        "歯科処置",
        "手術依頼",
        "手術実施",
        "手術材料等",
        "手術記録",
        "手術所見",
        "輸血",
        "輸血実施",
        "透析実施",
        "放射線治療実施",
        "リハビリ",
        "看護指示受け",
    }
)

ORDER_RECORD_TABLES = frozenset(
    {
        "病名",
        "サマリ管理",
        "サマリ病名",
        "サマリ記事",
        "インフォームドコンセント",
        "文書",
        "患者適用クリニカルパス",
        "患者適用クリニカルパス項目",
        "患者適用クリニカルパス記事",
    }
)


@dataclass(frozen=True)
class ColumnInfo:
    """カラム情報."""

    name: str
    sql_type: str
    default_value: str | int | float | date | time | None
    description: str | None


@dataclass
class TableInfo:
    """テーブル情報."""

    name: str
    group: str
    description: str
    record_unit: str
    columns: list[ColumnInfo] = field(default_factory=list)


def read_xlsx(path: Path) -> dict[str, TableInfo]:
    """xlsxファイルからテーブル情報を読み込む."""
    wb = openpyxl.load_workbook(path, read_only=True)

    # テーブル一覧シートからグループ・説明を取得
    tables: dict[str, TableInfo] = {}
    ws_tables = wb["テーブル一覧"]
    for row in ws_tables.iter_rows(min_row=2, values_only=True):
        group, table_name, desc, unit = row
        if table_name is not None:
            tables[table_name] = TableInfo(
                name=table_name,
                group=str(group or "OTHER"),
                description=str(desc or ""),
                record_unit=str(unit or ""),
            )

    # 項目一覧シートからカラム情報を取得
    ws_columns = wb["項目一覧"]
    for row in ws_columns.iter_rows(min_row=2, values_only=True):
        table_name, col_name, data_type, default_val, description, _update_date = row
        if table_name is None or col_name is None:
            continue

        # テーブル一覧に存在しないテーブルは OTHER として追加
        if table_name not in tables:
            tables[table_name] = TableInfo(
                name=table_name,
                group="OTHER",
                description="",
                record_unit="",
            )

        column = ColumnInfo(
            name=str(col_name),
            sql_type=str(data_type or "VARCHAR"),
            default_value=default_val,
            description=str(description) if description else None,
        )
        tables[table_name].columns.append(column)

    wb.close()
    return tables


def _sanitize_field_name(name: str) -> str:
    """フィールド名として使えない文字を置換する."""
    # 改行をアンダースコアに
    sanitized = name.replace("\n", "_").replace("\r", "_")
    # 先頭が数字の場合はプレフィックスを付ける
    if sanitized and sanitized[0].isdigit():
        sanitized = f"_{sanitized}"
    # スペースをアンダースコアに
    sanitized = sanitized.replace(" ", "_").replace("　", "_")
    # Pythonの識別子として無効な文字を除去（日本語・英数字・アンダースコア以外）
    sanitized = re.sub(r"[^\w]", "_", sanitized)
    return sanitized


def _format_default_value(sql_type: str, default_val: object) -> str | None:
    """デフォルト値をPythonリテラルに変換する. Noneなら返さない."""
    if default_val is None:
        return None

    python_type = SQL_TYPE_MAP.get(sql_type, "str")

    if python_type == "str":
        # 文字列型: シングルクォートを除去
        val = str(default_val).strip("'")
        return repr(val)

    if python_type == "int":
        try:
            return str(int(default_val))
        except (ValueError, TypeError):
            return None

    if python_type == "float":
        try:
            return str(float(default_val))
        except (ValueError, TypeError):
            return None

    if python_type == "Decimal":
        val = str(default_val).strip("'")
        return f'Decimal("{val}")'

    if python_type == "date":
        val = str(default_val).strip("'")
        # datetime.date型の場合
        if isinstance(default_val, date):
            return f"date({default_val.year}, {default_val.month}, {default_val.day})"
        # 文字列の場合 YYYY-MM-DD パース
        match = re.match(r"(\d{4})-(\d{2})-(\d{2})", val)
        if match:
            y, m, d = int(match.group(1)), int(match.group(2)), int(match.group(3))
            return f"date({y}, {m}, {d})"
        return None

    if python_type == "time":
        val = str(default_val).strip("'")
        # datetime.time型の場合
        if isinstance(default_val, time):
            return f"time({default_val.hour}, {default_val.minute}, {default_val.second})"
        # 文字列の場合 HH:MM:SS パース
        match = re.match(r"(\d{2}):(\d{2}):(\d{2})", val)
        if match:
            h, mi, s = int(match.group(1)), int(match.group(2)), int(match.group(3))
            return f"time({h}, {mi}, {s})"
        return None

    return None


def _resolve_duplicates(columns: list[ColumnInfo]) -> list[tuple[ColumnInfo, str]]:
    """重複するカラム名にサフィックスを付与する.

    Returns:
        (ColumnInfo, resolved_field_name) のリスト
    """
    name_count: dict[str, int] = {}
    result: list[tuple[ColumnInfo, str]] = []

    for col in columns:
        sanitized = _sanitize_field_name(col.name)
        name_count[sanitized] = name_count.get(sanitized, 0) + 1
        if name_count[sanitized] > 1:
            resolved = f"{sanitized}_{name_count[sanitized]}"
        else:
            resolved = sanitized
        result.append((col, resolved))

    # 最初の出現にも番号を付ける（重複がある場合のみ）
    final_counts: dict[str, int] = {}
    for col in columns:
        sanitized = _sanitize_field_name(col.name)
        final_counts[sanitized] = final_counts.get(sanitized, 0) + 1

    duplicated_names = {name for name, count in final_counts.items() if count > 1}

    # 再走査：最初の出現に _1 を付ける
    name_count2: dict[str, int] = {}
    result2: list[tuple[ColumnInfo, str]] = []
    for col in columns:
        sanitized = _sanitize_field_name(col.name)
        name_count2[sanitized] = name_count2.get(sanitized, 0) + 1
        if sanitized in duplicated_names:
            resolved = f"{sanitized}_{name_count2[sanitized]}"
        else:
            resolved = sanitized
        result2.append((col, resolved))

    return result2


def _table_name_to_class_name(table_name: str) -> str:
    """テーブル名をクラス名に変換する."""
    # 記号をアンダースコアに置換
    name = table_name.replace("・", "_").replace("(", "_").replace(")", "")
    name = name.replace("（", "_").replace("）", "")
    name = re.sub(r"[^\w]", "_", name)
    # 連続アンダースコアを1つに
    name = re.sub(r"_+", "_", name).strip("_")
    return name


def generate_model_code(table: TableInfo) -> str:
    """1テーブル分のPydanticモデルコードを生成する."""
    class_name = _table_name_to_class_name(table.name)
    lines: list[str] = []

    # docstring
    doc = table.description or table.name
    if table.record_unit:
        doc += f"（{table.record_unit}）"
    lines.append(f'class {class_name}(DwhBaseModel):')
    lines.append(f'    """{doc}."""')
    lines.append("")

    resolved_columns = _resolve_duplicates(table.columns)

    if not resolved_columns:
        lines.append("    pass")
        return "\n".join(lines)

    for col, field_name in resolved_columns:
        python_type = SQL_TYPE_MAP.get(col.sql_type, "str")
        default_literal = _format_default_value(col.sql_type, col.default_value)
        needs_alias = field_name != _sanitize_field_name(col.name)

        if default_literal is not None:
            # デフォルト値あり → nullable ではない
            if needs_alias:
                lines.append(
                    f"    {field_name}: {python_type}"
                    f' = Field({default_literal}, alias="{col.name}")'
                )
            else:
                lines.append(f"    {field_name}: {python_type} = {default_literal}")
        else:
            # デフォルト値なし → nullable
            if needs_alias:
                lines.append(
                    f"    {field_name}: {python_type} | None"
                    f' = Field(None, alias="{col.name}")'
                )
            else:
                lines.append(f"    {field_name}: {python_type} | None = None")

    return "\n".join(lines)


def _collect_imports(tables: list[TableInfo]) -> set[str]:
    """テーブル群から必要なimportを収集する."""
    imports: set[str] = set()
    needs_field = False

    for table in tables:
        resolved = _resolve_duplicates(table.columns)
        for col, field_name in resolved:
            sql_type = col.sql_type
            if sql_type == "DATE":
                imports.add("date")
            elif sql_type == "TIME":
                imports.add("time")
            elif sql_type == "DECIMAL":
                imports.add("Decimal")
            if field_name != _sanitize_field_name(col.name):
                needs_field = True

    return imports | ({"Field"} if needs_field else set())


def _build_import_block(type_imports: set[str]) -> str:
    """import文ブロックを組み立てる."""
    lines: list[str] = []
    lines.append('"""自動生成されたDWHテーブルモデル."""')
    lines.append("")
    lines.append("from __future__ import annotations")
    lines.append("")

    # datetime imports
    datetime_types = sorted(type_imports & {"date", "time"})
    if datetime_types:
        lines.append(f"from datetime import {', '.join(datetime_types)}")

    # decimal imports
    if "Decimal" in type_imports:
        lines.append("from decimal import Decimal")

    # pydantic imports
    pydantic_imports = ["Field"] if "Field" in type_imports else []
    if pydantic_imports:
        lines.append(f"from pydantic import {', '.join(sorted(pydantic_imports))}")

    lines.append("")
    lines.append("from interactive_ehr.models._base import DwhBaseModel")

    return "\n".join(lines)


def _get_file_key(table: TableInfo) -> str:
    """テーブルの出力先ファイルキーを決定する."""
    group = table.group

    if group == "ORDER" or table.name in ORDER_EXAM_TABLES | ORDER_TREATMENT_TABLES | ORDER_RECORD_TABLES:
        if table.name in ORDER_EXAM_TABLES:
            return "order_exam"
        if table.name in ORDER_TREATMENT_TABLES:
            return "order_treatment"
        if table.name in ORDER_RECORD_TABLES:
            return "order_record"
        # ORDER グループだがどのサブにも属さない → exam にフォールバック
        return "order_exam"

    return GROUP_FILE_MAP.get(group, "other")


def generate_all(tables: dict[str, TableInfo]) -> dict[str, str]:
    """全テーブルをファイルごとにグループ化してコードを生成する."""
    file_tables: dict[str, list[TableInfo]] = {}

    for table in tables.values():
        key = _get_file_key(table)
        if key not in file_tables:
            file_tables[key] = []
        file_tables[key].append(table)

    generated_files: dict[str, str] = {}

    for file_key, table_list in sorted(file_tables.items()):
        type_imports = _collect_imports(table_list)
        import_block = _build_import_block(type_imports)

        model_blocks: list[str] = []
        for table in table_list:
            model_blocks.append(generate_model_code(table))

        content = import_block + "\n\n\n" + "\n\n\n".join(model_blocks) + "\n"
        generated_files[file_key] = content

    return generated_files


def generate_init_py(tables: dict[str, TableInfo]) -> str:
    """__init__.py を生成する."""
    lines: list[str] = []
    lines.append('"""DWHテーブルモデル."""')
    lines.append("")
    lines.append("from interactive_ehr.models._base import DwhBaseModel")
    lines.append("")

    file_classes: dict[str, list[str]] = {}
    for table in tables.values():
        file_key = _get_file_key(table)
        class_name = _table_name_to_class_name(table.name)
        if file_key not in file_classes:
            file_classes[file_key] = []
        file_classes[file_key].append(class_name)

    for file_key in sorted(file_classes):
        classes = sorted(file_classes[file_key])
        for cls in classes:
            lines.append(
                f"from interactive_ehr.models.{file_key} import {cls}"
            )

    lines.append("")
    all_classes = ["DwhBaseModel"]
    for file_key in sorted(file_classes):
        all_classes.extend(sorted(file_classes[file_key]))

    lines.append("__all__ = [")
    for cls in all_classes:
        lines.append(f'    "{cls}",')
    lines.append("]")
    lines.append("")

    return "\n".join(lines)


def main() -> None:
    """メイン処理."""
    if not XLSX_PATH.exists():
        print(f"Error: {XLSX_PATH} が見つかりません", file=sys.stderr)
        sys.exit(1)

    print(f"読み込み中: {XLSX_PATH}")
    tables = read_xlsx(XLSX_PATH)
    print(f"テーブル数: {len(tables)}")

    total_columns = sum(len(t.columns) for t in tables.values())
    print(f"総カラム数: {total_columns}")

    # モデルファイル生成
    generated = generate_all(tables)
    for file_key, content in generated.items():
        output_path = MODELS_DIR / f"{file_key}.py"
        output_path.write_text(content, encoding="utf-8")
        print(f"生成: {output_path}")

    # __init__.py 生成
    init_content = generate_init_py(tables)
    init_path = MODELS_DIR / "__init__.py"
    init_path.write_text(init_content, encoding="utf-8")
    print(f"生成: {init_path}")

    # ruff format
    print("ruff format 実行中...")
    subprocess.run(
        ["uv", "run", "ruff", "format", str(MODELS_DIR)],
        check=True,
        cwd=PROJECT_ROOT,
    )
    print("完了!")


if __name__ == "__main__":
    main()
